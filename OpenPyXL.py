from openpyxl.styles import PatternFill, Side, Border
from Alert import say, alert
from time import sleep


def is_workbook_blocked(workbook, dest):
    try:
        workbook.save(dest)
    except PermissionError:
        return True
    return False


def safe_worbook_save(workbook, dest):
    while is_workbook_blocked(workbook, dest):
        say("Workbook is open, close it")
        print("Workbook is open, close it")
        sleep(1)
    pass


def cells_style(sheet, x_max=100, y_max=100, colored=False):
    def cell_default_style(cell):
        thin_border = Border(left=Side(style='thin', color='00000000'),
                             right=Side(style='thin', color='00000000'),
                             top=Side(style='thin', color='00000000'),
                             bottom=Side(style='thin', color='FFFFFFFF'))
        if colored:
            cell.fill = PatternFill(start_color="3A3838", end_color="3A3838", fill_type="solid")
        cell.border = thin_border
        if cell.parent.column_dimensions[cell.column_letter].width < len(str(cell.value)):
            cell.parent.column_dimensions[cell.column_letter].width = len(str(cell.value))

    for y in range(1, y_max + 5 + 1):
        for x in range(1, max(5, x_max + 5 + 1)):
            cell_default_style(sheet.cell(y, x))
    sheet.sheet_view.zoomScale = 140
